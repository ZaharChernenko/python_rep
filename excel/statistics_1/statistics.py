from dataclasses import dataclass
from re import split

import pandas as pd
from numpy import ndarray
from openpyxl.chart import Reference, ScatterChart, Series
from openpyxl.styles import PatternFill


@dataclass
class DiscreteRandomVars:
    row_name: str
    col_name: str

    row_values: tuple[float]
    col_values: tuple[float]

    probability_matrix: ndarray

    expected_val4row: float
    expected_val4col: float
    dispersion4row: float
    dispersion4col: float
    stand_deviation4row: float
    stand_deviation4col: float

    covariance: float
    correlation: float

    def __init__(self, sheet, name: str):
        self.row_name, self.col_name = self._readVarsNames(sheet)
        self.row_values = self._readRowsValues(sheet)
        self.col_values = self._readColValues(sheet)

        self.probability_matrix = pd.read_excel(name, index_col=0).to_numpy()
        self._checkProb()

        self.expected_val4row = self._getExpVal4Row()
        self.expected_val4col = self._getExpVal4Col()
        self.dispersion4row = self._getDispersion4Row()
        self.dispersion4col = self._getDispersion4Col()
        self.stand_deviation4row = round(self.dispersion4row ** 0.5, 2)
        self.stand_deviation4col = round(self.dispersion4col ** 0.5, 2)

        self.covariance = self._getCovariance()
        self.correlation = self._getCorrelation()

    def _readVarsNames(self, sheet) -> list[str]:
        return split(r"\W+", sheet["A1"].value)

    def _readRowsValues(self, sheet) -> tuple[int | float]:
        return tuple(*sheet.iter_cols(min_row=2, max_row=sheet.max_row, max_col=1, values_only=True))

    def _readColValues(self, sheet) -> tuple[int | float]:
        return tuple(*sheet.iter_rows(max_row=1, min_col=2, max_col=sheet.max_column, values_only=True))

    def _checkProb(self) -> None:
        """Raises exception if total probability != 1"""
        total_prob = 0
        for row in self.probability_matrix:
            for elem in row:
                total_prob += elem
        if round(total_prob, 3) != 1:
            raise ArithmeticError("Суммарная вероятность не равна 1!")

    def _getExpVal4Row(self) -> float:
        expected_val4row: float = 0

        for i in range(len(self.row_values)):
            row_probability = 0
            for j in range(len(self.probability_matrix[i])):
                row_probability += self.probability_matrix[i][j]
            expected_val4row += row_probability * self.row_values[i]

        return round(expected_val4row, 2)

    def _getExpVal4Col(self) -> float:
        expected_val4col: float = 0

        for j in range(len(self.col_values)):
            col_probability = 0
            for i in range(len(self.probability_matrix)):
                col_probability += self.probability_matrix[i][j]
            expected_val4col += col_probability * self.col_values[j]

        return round(expected_val4col, 2)

    def _getDispersion4Row(self):
        dispersion4row: float = 0

        for i in range(len(self.row_values)):
            row_probability = 0
            for j in range(len(self.probability_matrix[i])):
                row_probability += self.probability_matrix[i][j]
            dispersion4row += row_probability * self.row_values[i] ** 2

        return round(dispersion4row - self.expected_val4row ** 2, 2)

    def _getDispersion4Col(self):
        dispersion4col: float = 0

        for j in range(len(self.col_values)):
            col_probability = 0
            for i in range(len(self.probability_matrix)):
                col_probability += self.probability_matrix[i][j]
            dispersion4col += col_probability * self.col_values[j] ** 2

        return round(dispersion4col - self.expected_val4col ** 2, 2)

    def _getCovariance(self) -> float:
        expected_val_mul = 0
        for i in range(len(self.row_values)):
            for j in range(len(self.col_values)):
                expected_val_mul += self.probability_matrix[i][j] * (self.row_values[i] * self.col_values[j])

        return round(expected_val_mul - self.expected_val4row * self.expected_val4col, 2)

    def _getCorrelation(self) -> float:
        return round(self.covariance / self.stand_deviation4row / self.stand_deviation4col, 2)

    def writeBaseValues(self, sheet):
        sheet["A1"].value = f"M({self.col_name})= {self.expected_val4col}"
        sheet["B1"].value = f"D({self.col_name})= {self.dispersion4col}"
        sheet["C1"].value = f"σ({self.col_name})= {self.stand_deviation4col}"

        sheet["A2"].value = f"М({self.row_name})= {self.expected_val4row}"
        sheet["B2"].value = f"D({self.row_name})= {self.dispersion4row}"
        sheet["C2"].value = f"σ({self.row_name})= {self.stand_deviation4row}"

        sheet["D1"].value = f"C({self.col_name}, {self.row_name})= {self.covariance}"
        sheet["D2"].value = f"r({self.col_name}, {self.row_name})= {self.correlation}"

        quantities_arr = ["A1", "B1", "C1", "A2", "B2", "C2", "D1", "D2"]
        for cell in quantities_arr:
            sheet[cell].fill = PatternFill(patternType="solid", fgColor="8497b0")

    def writeRowExpValSubject2ColVal(self, sheet):
        col_prob_arr = []
        for j in range(len(self.col_values)):
            prob = 0
            for i in range(len(self.row_values)):
                prob += self.probability_matrix[i][j]
            col_prob_arr.append(round(prob, 3))

        graphic: list[tuple[float, float]] = []
        start_row = 4
        for j in range(len(self.col_values)):
            sheet[f"A{start_row}"].value = f"{self.col_name} = {self.col_values[j]}"
            start_row += 1
            sheet[f"A{start_row}"].value = self.row_name
            start_row += 1
            sheet[f"A{start_row}"].value = f"P({self.row_name}|{sheet[f"A{start_row - 2}"].value})"

            for cell in range(start_row - 2, start_row + 1):
                sheet[f"A{cell}"].fill = PatternFill(patternType="solid", fgColor="8497b0")

            start_column = 66
            expected_val4row_subject2col = 0
            for i in range(len(self.row_values)):
                sheet[f"{chr(start_column)}{start_row - 1}"].value = self.row_values[i]
                sheet[f"{chr(start_column)}{start_row}"].value = round(
                    self.probability_matrix[i][j] / col_prob_arr[j], 3)
                start_column += 1
                expected_val4row_subject2col += self.probability_matrix[i][j] / col_prob_arr[j] * self.row_values[i]
            graphic.append((self.col_values[j], round(expected_val4row_subject2col, 2)))
            sheet[f"{chr(start_column)}{start_row}"].value = (f"M({self.row_name}|{self.col_name} = "
                                                              f"{self.col_values[j]}) = {graphic[j][1]}")
            sheet[f"{chr(start_column)}{start_row}"].fill = PatternFill(patternType="solid", fgColor="8497b0")
            start_row += 2

        return graphic, start_row + 1

    def writeGraphic(self, sheet, graphic: list[tuple[float, float]], row: int):
        sheet[f"A{row}"].value = f"{self.col_name}"
        sheet[f"B{row}"].value = f"{self.row_name} = Reg({self.col_name})"
        sheet[f"C{row}"].value = f"{self.row_name} = lin Reg({self.col_name})"
        row += 1
        start_row = row

        for j in range(len(self.col_values)):
            sheet[f"A{row}"].value = graphic[j][0]
            sheet[f"B{row}"].value = graphic[j][1]
            sheet[f"C{row}"].value = round(self.correlation * (self.stand_deviation4row / self.stand_deviation4col) *
                                           (self.col_values[j] - self.expected_val4col) + self.expected_val4row, 2)
            row += 1

        chart = ScatterChart()
        chart.title = "Графики истинной и среднеквадратической линейной регрессий"
        chart.x_axis.title = self.col_name
        chart.y_axis.title = self.row_name
        xvalues = Reference(sheet, min_row=start_row, max_row=row, min_col=1)

        yvalues1 = Reference(sheet, min_row=start_row - 1, max_row=row, min_col=2)
        series = Series(yvalues1, xvalues, title_from_data=True)
        chart.series.append(series)

        yvalues2 = Reference(sheet, min_row=start_row - 1, max_row=row, min_col=3)
        series = Series(yvalues2, xvalues, title_from_data=True)
        chart.series.append(series)

        sheet.add_chart(chart, f"A{row}")
