import openpyxl


class ExcelReader:
    def __init__(self, file_path: str):
        workbook = openpyxl.load_workbook(file_path)
        self._worksheet = workbook.active
        # считываем первый ряд
        self._headers: tuple[str] = tuple(
            filter(
                lambda el: el is not None,
                *self._worksheet.iter_rows(max_row=1, min_col=1, max_col=self._worksheet.max_column, values_only=True)
            )
        )
        self._len: int = 0
        for step, row in enumerate(
            self._worksheet.iter_rows(
                min_row=2, max_row=self._worksheet.max_row, max_col=len(self._headers), values_only=True
            ),
            start=2,
        ):
            if not any(row):
                break
            self._len = step

    def __iter__(self):
        for row in self._worksheet.iter_rows(
            min_row=2, max_row=self._len, max_col=len(self._headers), values_only=True
        ):
            yield {self._headers[i]: row[i] for i in range(len(self._headers))}

    def __len__(self):
        return self._len


if __name__ == "__main__":
    test_reader: ExcelReader = ExcelReader("tests/table.xlsx")
    for el in test_reader:
        print(el)
